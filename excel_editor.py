from openpyxl import load_workbook
import json

def get_name_dict():
    with open('gene_names.json') as json_file:
        names_data = json.load(json_file)
        return names_data

def change_gene_names(names_data):
    #load excel file
    wb = load_workbook(filename="MAGMA_Gene_Level_Analysis_Results_12_01_22.xlsx")

    for sheet in wb:
        print(sheet.title)
        #no cells to change in first sheet
        if sheet.title == "Overview":
            continue
        sheet.insert_cols(1)
        sheet["A1"] = "Name"
        #modify row
        for a,b in sheet.iter_rows(min_row=2, min_col=1, max_col=2):
            name = names_data[str(b.value)]
            print(name)
            a.value = name
    wb.save("MAGMA_Gene_Level_Analysis_Results_12_01_22_Modified.xlsx")
    print("Done")



if __name__ == "__main__":
    names_data = get_name_dict()
    change_gene_names(names_data)